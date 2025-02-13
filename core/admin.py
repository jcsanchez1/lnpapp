from django.contrib import admin
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.contrib import admin
from .models import Expediente, Region, CentroAtencion, Profile, Muestra


@admin.register(Expediente)
class ExpedienteAdmin(admin.ModelAdmin):
    list_display = ('dni', 'nombre', 'apellido', 'sexo', 'fecha_creacion', 'fecha_modificacion')
    search_fields = ('dni', 'nombre', 'apellido')
    list_filter = ('sexo', 'fecha_creacion', 'fecha_modificacion')


@admin.register(Region)
class RegionAdmin(admin.ModelAdmin):
    list_display = ('numero_region', 'nombre')
    search_fields = ('nombre',)
    list_filter = ('numero_region',)


@admin.register(CentroAtencion)
class CentroAtencionAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'direccion', 'region')
    search_fields = ('nombre', 'direccion')
    list_filter = ('region',)


@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'centro_atencion')
    search_fields = ('user__username', 'centro_atencion__nombre')
    list_filter = ('centro_atencion',)


class MuestraAdmin(admin.ModelAdmin):
    fieldsets = [
        ('Información General', {
            'fields': ['CentroAtencion', 'Expediente', 'Edad'],
            'classes': ['tab-general']
        }),
        ('Características de la Muestra', {
            'fields': ['consistencia', 'moco', 'sangre_macroscopica'],
            'classes': ['tab-caracteristicas']
        }),
        ('Protozoos - Amebas', {
            'fields': [
                'Entamoeba_histolytica',
                'Entamoeba_coli',
                'Entamoeba_hartmanni',
                'Endolimax_nana',
                'Iodamoeba_bütschlii'
            ],
            'classes': ['tab-amebas']
        }),
        ('Protozoos - Flagelados y Otros', {
            'fields': [
                'Blastocystis_sp',
                'Giardia_intestinalis',
                'Pentatrichomonas_hominis',
                'Chilomastix_mesnili',
                'Balantidium_coli'
            ],
            'classes': ['tab-flagelados']
        }),
        ('Protozoos - Coccidios', {
            'fields': [
                'Cystoisospora_belli',
                'Cyclospora_cayetanensis',
                'Criptosporidium_spp'
            ],
            'classes': ['tab-coccidios']
        }),
        ('Helmintos - Nematodos', {
            'fields': [
                'Strongyloides_stercoralis',
                'Ascaris_lumbricoides',
                'Trichuris_trichiura',
                'Necator_americanus',
                'Enterobius_vermicularis'
            ],
            'classes': ['tab-nematodos']
        }),
        ('Helmintos - Cestodos', {
            'fields': [
                'Taenia_spp',
                'Hymenolepis_diminuta',
                'Rodentolepis_nana'
            ],
            'classes': ['tab-cestodos']
        }),
        ('Resultados Adicionales', {
            'fields': [
                'Intensidad_de_la_Infección_KATO_KATZ',
                'No_se_observaron_parásitos'
            ],
            'classes': ['tab-resultados']
        }),
    ]

    list_display = [
        'fecha',
        'get_paciente_nombre',
        'get_paciente_dni',
        'Edad',
        'consistencia',
        'get_parasitos_encontrados'
    ]

    list_filter = [
        'fecha',
        'consistencia',
        'No_se_observaron_parásitos',
        'CentroAtencion'
    ]

    search_fields = [
        'Expediente__nombre',
        'Expediente__dni',
    ]

    readonly_fields = ['fecha']
#=======================================================================================================
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.export_excel_view, name='export_excel'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    def export_excel_view(self, request):
        if request.method == "POST":
            # Obtener fechas del formulario
            fecha_inicio = datetime.strptime(request.POST.get('fecha_inicio'), '%Y-%m-%d')
            fecha_fin = datetime.strptime(request.POST.get('fecha_fin'), '%Y-%m-%d')
            
            # Filtrar muestras por rango de fechas
            muestras = Muestra.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin
            ).select_related('Expediente', 'CentroAtencion')

            return self.generate_excel(muestras)
        
        # Si es GET, mostrar el formulario
        context = {
            'title': 'Exportar a Excel',
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/export_excel_form.html', context)

    def generate_excel(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Muestras"

        # Definir estilos
        header_style = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Cabeceras
        headers = [
            'Fecha',
            'DNI',
            'Nombre',
            'Apellido',
            'Edad',
            'Sexo',
            'Centro de Atención',
            'Consistencia',
            'Moco',
            'Sangre Macroscópica',
            'Entamoeba histolytica',
            'Entamoeba coli',
            'Entamoeba hartmanni',
            'Endolimax nana',
            'Iodamoeba bütschlii',
            'Blastocystis sp',
            'Giardia intestinalis',
            'Pentatrichomonas hominis',
            'Chilomastix mesnili',
            'Balantidium coli',
            'Cystoisospora belli',
            'Cyclospora cayetanensis',
            'Criptosporidium spp',
            'Strongyloides stercoralis',
            'Ascaris lumbricoides',
            'Trichuris trichiura',
            'Necator americanus',
            'Enterobius vermicularis',
            'Taenia spp',
            'Hymenolepis diminuta',
            'Rodentolepis nana',
            'Intensidad de la Infección KATO-KATZ',
            'No se observaron parásitos'
        ]

        # Escribir cabeceras
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style
            cell.fill = header_fill

        # Escribir datos
        for row, muestra in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=muestra.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=muestra.Expediente.dni)
            ws.cell(row=row, column=3, value=muestra.Expediente.nombre)
            ws.cell(row=row, column=4, value=muestra.Expediente.apellido)
            ws.cell(row=row, column=5, value=muestra.Edad)
            ws.cell(row=row, column=6, value=muestra.Expediente.get_sexo_display())
            ws.cell(row=row, column=7, value=muestra.CentroAtencion.nombre)
            ws.cell(row=row, column=8, value=muestra.get_consistencia_display())
            ws.cell(row=row, column=9, value=muestra.get_moco_display())
            ws.cell(row=row, column=10, value=muestra.get_sangre_macroscopica_display())
            ws.cell(row=row, column=11, value=muestra.get_Entamoeba_histolytica_display())
            ws.cell(row=row, column=12, value=muestra.get_Entamoeba_coli_display())
            ws.cell(row=row, column=13, value=muestra.get_Entamoeba_hartmanni_display())
            ws.cell(row=row, column=14, value=muestra.get_Endolimax_nana_display())
            ws.cell(row=row, column=15, value=muestra.get_Iodamoeba_bütschlii_display())
            ws.cell(row=row, column=16, value=muestra.get_Blastocystis_sp_display())
            ws.cell(row=row, column=17, value=muestra.get_Giardia_intestinalis_display())
            ws.cell(row=row, column=18, value=muestra.get_Pentatrichomonas_hominis_display())
            ws.cell(row=row, column=19, value=muestra.get_Chilomastix_mesnili_display())
            ws.cell(row=row, column=20, value=muestra.get_Balantidium_coli_display())
            ws.cell(row=row, column=21, value=muestra.get_Cystoisospora_belli_display())
            ws.cell(row=row, column=22, value=muestra.get_Cyclospora_cayetanensis_display())
            ws.cell(row=row, column=23, value=muestra.get_Criptosporidium_spp_display())
            ws.cell(row=row, column=24, value=muestra.get_Strongyloides_stercoralis_display())
            ws.cell(row=row, column=25, value=muestra.get_Ascaris_lumbricoides_display())
            ws.cell(row=row, column=26, value=muestra.get_Trichuris_trichiura_display())
            ws.cell(row=row, column=27, value=muestra.get_Necator_americanus_display())
            ws.cell(row=row, column=28, value=muestra.get_Enterobius_vermicularis_display())
            ws.cell(row=row, column=29, value=muestra.get_Taenia_spp_display())
            ws.cell(row=row, column=30, value=muestra.get_Hymenolepis_diminuta_display())
            ws.cell(row=row, column=31, value=muestra.get_Rodentolepis_nana_display())
            ws.cell(row=row, column=32, value=muestra.get_Intensidad_de_la_Infección_KATO_KATZ_display())
            ws.cell(row=row, column=33, value='Sí' if muestra.No_se_observaron_parásitos else 'No')

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Crear respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=reporte_muestras.xlsx'
        wb.save(response)
        return response
#=======================================================================================================
    def get_paciente_nombre(self, obj):
        return obj.Expediente.nombre
    get_paciente_nombre.short_description = 'Nombre del Paciente'

    def get_paciente_dni(self, obj):
        return obj.Expediente.dni
    get_paciente_dni.short_description = 'DNI'

    def get_parasitos_encontrados(self, obj):
        parasitos = []
        campos_parasitos = [
            'Entamoeba_histolytica', 'Entamoeba_coli', 'Giardia_intestinalis',
            'Blastocystis_sp', 'Ascaris_lumbricoides', 'Trichuris_trichiura'
        ]
        
        for campo in campos_parasitos:
            valor = getattr(obj, campo)
            if valor != 'N':  # Si no es "No se observa"
                parasitos.append(campo.replace('_', ' '))
        
        return ', '.join(parasitos) if parasitos else 'Ninguno'
    get_parasitos_encontrados.short_description = 'Parásitos Encontrados'

    class Media:
        css = {
            'all': ('admin/css/custom_tabs.css',)
        }
        js = ('admin/js/custom_tabs.js',)

admin.site.register(Muestra, MuestraAdmin)
